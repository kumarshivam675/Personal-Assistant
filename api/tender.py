from openpyxl import load_workbook


def getWard(location):
    currx = float(location[0])
    curry = float(location[1])
    print location
    wb = load_workbook('/home/shivam/whatsapp/testing/janagraaha-test/data/ward.xlsx')
    sheet = wb.get_sheet_by_name('Sheet3')
    ward = sheet.cell(row=1, column=1).value
    x = float(sheet.cell(row=1, column=4).value)
    y = float(sheet.cell(row=1, column=5).value)
    minDistance = (currx - x) * (currx - x) + (curry - y) * (curry - y)
    for i in range(2, 171):
        x = float(sheet.cell(row=i, column=4).value)
        y = float(sheet.cell(row=i, column=5).value)
        distance = (currx - x) * (currx - x) + (curry - y) * (curry - y)
        if distance <= minDistance:
            minDistance = distance
            ward = str(sheet.cell(row=i, column=1).value)

    return ward


def getTenders(ward):
    tenders = ""
    count = 0
    print ward
    wb = load_workbook('/home/shivam/whatsapp/testing/janagraaha-test/data/ward.xlsx')
    sheet = wb.get_sheet_by_name('BBMP Tenders')
    for i in range(2, 3500):
        wardName = sheet.cell(row=i, column=3).value
        if wardName == ward:
            count += 1
            if count > 4:
                return tenders
            ten = "Sl No: " + str(sheet.cell(row=i, column=1).value) + " Title: " + sheet.cell(row=i,
                                                                                               column=12).value + " Estimated Value: Rs" + str(
                sheet.cell(row=i, column=16).value) + " End Date: " + str(sheet.cell(row=i, column=9).value)
            tenders += "\n\n" + ten
    return tenders


def getComplains(ward):
    ward = ward.replace(" ", "")
    ward = ward.lower()
    complains = ""
    count = 0
    wb = load_workbook('/home/shivam/whatsapp/testing/janagraaha-test/data/ICMyC_complaints.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(2, 1000):
        wardName = sheet.cell(row=i, column=5).value
        wardName = wardName.replace(" ", "")
        wardName = wardName.lower()
        if wardName == ward:
            count += 1
            if count >= 3:
                return complains
            ten = "Complaint No: " + sheet.cell(row=i, column=2).value + "\nComplaint Title: " + sheet.cell(row=i,
                                                                                                            column=6).value + "\nLocation:" + sheet.cell(
                row=i, column=10).value + "\n Status: " + sheet.cell(row=i, column=14).value
            complains += "\n\n" + ten

    return complains

# location = [12.8500, 77.6700]
#
# # print getWard(location)
# ward = "Singa Sandra"
# tenders = getTenders(ward)
# for i in tenders:
#     print i
