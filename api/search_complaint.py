from openpyxl import load_workbook

def search(id):
    wb = load_workbook('/home/shivam/whatsapp/testing/janagraaha-test/data/ICMyC_complaints.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    details = ""
    for i in range(2,400):
        print "hello"
        inputId = sheet.cell(row=i, column=2).value
        print inputId
        if inputId == id:
            details += "Title:"
            details += sheet.cell(row=i, column=6).value
            details += ".\nComplaint was filed on "
            details += str(sheet.cell(row=i, column=3).value)
            details += ".\nThe status of your complaint is "
            details += sheet.cell(row=i, column=14).value
            print details
            return details
    return details

# if __name__ == "__main__":
#     print "hi"
#     search("W15011008")